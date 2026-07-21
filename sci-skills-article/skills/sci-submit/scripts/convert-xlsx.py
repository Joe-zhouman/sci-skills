#!/usr/bin/env python3
"""Convert 高质量科技期刊分级目录速查表.xlsx to a searchable JSON index.

Usage:
  python3 convert-xlsx.py <xlsx_path> [output_json_path]

Output JSON structure:
  [
    {
      "journal": "Nature Energy",
      "field": "能源电力领域",
      "tier": "T1",
      "issn": "2058-7546",
      "jcr": "Q1",
      "cas_base": "工程技术1区",
      "cas_upgrade": "材料科学1区",
      "cas_small": "能源与燃料1区/材料科学：综合1区",
      "cas_top": "材料科学TOP",
      "if": "70.1",
      "if5": "66.6",
      "cas_warn": null,
      "indexed_by": "SCI, ESI, EI"
    },
    ...
  ]

Run this whenever the xlsx source file is updated.
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Please install openpyxl: pip install openpyxl")

# Column name normalization map — unify varying headers across sheets
COLUMN_MAP = {
    "期刊名": "journal",
    "分级": "tier",
    "ISSN": "issn",
    "ISSN/EISSN": "issn",
    "CN": "cn",
    "学科领域": "subfield",
    "JCR分区": "jcr",
    "中科院基础版": "cas_base",
    "中科院升级版": "cas_upgrade",
    "中科院小类": "cas_small",
    "中科院Top": "cas_top",
    "影响因子": "if_val",
    "五年影响因子": "if5",
    "中科院预警": "cas_warn",
    "检索库": "indexed_by",
    "序号": "row_num",
}


def normalize_value(val):
    """Convert openpyxl cell value to a JSON-safe value."""
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() == "none":
        return None
    return s


def convert_sheet(ws, field_name):
    """Convert a single sheet to a list of dicts."""
    headers = []
    for c in range(1, ws.max_column + 1):
        raw = str(ws.cell(row=1, column=c).value or "").strip()
        headers.append(COLUMN_MAP.get(raw, raw.lower().replace(" ", "_")))

    records = []
    for r in range(2, ws.max_row + 1):
        record = {"field": field_name}
        for c, key in enumerate(headers, start=1):
            val = normalize_value(ws.cell(row=r, column=c).value)
            if val is not None:
                record[key] = val
        # Skip empty rows
        if record.get("journal"):
            records.append(record)
    return records


BUNDLED_XLSX = Path(__file__).resolve().parent.parent / "data" / "高质量科技期刊分级目录速查表.xlsx"
BUNDLED_OUT = Path(__file__).resolve().parent.parent / "data" / "journal-ratings.json"


def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else BUNDLED_XLSX
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else BUNDLED_OUT

    if not xlsx_path.exists():
        sys.exit(f"xlsx not found: {xlsx_path}\n"
                 f"Usage: {sys.argv[0]} [xlsx_path] [output_json_path]\n"
                 f"Defaults to bundled copy in assets/.")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    all_records = []
    # Skip sheet 0 (目录索引), process all data sheets
    for name in wb.sheetnames[1:]:
        ws = wb[name]
        all_records.extend(convert_sheet(ws, name))

    wb.close()

    output_path.write_text(
        json.dumps(all_records, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"Wrote {len(all_records)} journal records from {len(wb.sheetnames) - 1} fields to {output_path}")


if __name__ == "__main__":
    main()
